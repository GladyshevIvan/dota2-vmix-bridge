import os
from pandas import DataFrame, ExcelWriter
from wx import MessageBox, OK
from openpyxl import load_workbook
from app.database_manager import Dota2TournamentsDataBase
from app.database_requests import show_match_stats


def make_datasource_excel(teams_df, players_df, path=''):
    '''
        Создает или обновляет Excel-файл с данными о командах и игроках.

        Args:
            teams_df (pandas.DataFrame): Данные о командах.
            players_df (pandas.DataFrame): Данные об игроках.
            path (str): Путь для сохранения Excel-файла (по умолчанию '').
    '''

    try:
        excel_file = os.path.join(path, 'DOTA 2 Stats.xlsx')
        sheet_name = 'DOTA 2'

        if os.path.isfile(excel_file):
            #Если файл уже создан, меняется только страница 'DOTA 2', на которой перезаписываются обе таблицы
            file = load_workbook(excel_file)

            if sheet_name in file.sheetnames:
                sheet = file[sheet_name]
            else:
                sheet = file.create_sheet(title=sheet_name)

            with ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                teams_df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    startrow=0,
                    startcol=0,
                )

            with ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                players_df.to_excel(
                    writer,
                    sheet_name=sheet_name,
                    index=False,
                    startrow=len(teams_df) + 2,
                    startcol=0,
                )

        else:
            #Если файл не существует, в новы файл на страницу 'DOTA 2' записываются две таблицы
            with ExcelWriter(excel_file, engine='xlsxwriter') as writer:
                teams_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0, startcol=0)
                players_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(teams_df) + 2, startcol=0)

        MessageBox('Excel-файл для Vmix создан', caption='Внимание', style=OK)
    except PermissionError as err:
        MessageBox('Пожалуйста, закройте Excel-файл и повторите попытку', caption='Внимание', style=OK)


def convert_stats_from_api(data, path):
    '''
        Конвертирует полученную из API статистику в Excel-файл

        Args:
            data (dict): статистика матча, внутри два словаря "teams_stats" и "players_stats"
            path (str): Путь для сохранения Excel-файла
    '''


    teams_stats = data['teams_stats']
    players_stats = data['players_stats']

    teams_df = DataFrame.from_dict(teams_stats, orient='index')
    teams_df.index.name = 'side'

    players_df = DataFrame.from_dict(players_stats, orient='index')
    players_df.index.name = 'id'

    make_datasource_excel(teams_df, players_df, path)


def convert_stats_from_db(match_id, path, database_name):
    '''
            Конвертирует полученную из БД статистику в Excel-файл

            Args:
                match_id (int): ID матча для получения статистики
                path (str): Путь для сохранения Excel-файла
                database_name (str): Имя базы данных
        '''

    db = Dota2TournamentsDataBase(database_name)
    data = show_match_stats(db, match_id)

    teams_stats = data['teams_stats']
    players_stats = data['players_stats']

    #Проверка была ли в data статистика матча, если нет, значит матча с таким ID нет в Базе Данных
    if teams_stats and players_stats:
        teams_df = DataFrame.from_dict(teams_stats)
        players_df = DataFrame.from_dict(players_stats)
        make_datasource_excel(teams_df, players_df, path)
    else:
        MessageBox(f'Матча с ID: {match_id} нет в Базе Данных {database_name}')
